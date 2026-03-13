#!/usr/bin/env python3
"""
Script to parse the Excel file "Jeux et mécaniques.xlsx" and extract:
1. Categories and Mechanics definitions
2. Categories for each game
3. Mechanics for each game

Usage:
    python scripts/parse-excel.py /path/to/Jeux\ et\ mécaniques.xlsx

Requirements:
    pip install openpyxl
"""

import sys
import json
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def parse_excel(excel_path):
    """Parse the Excel file and extract all data."""
    
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    result = {
        "categories_definitions": {},
        "mechanics_definitions": {},
        "game_categories": {},
        "game_mechanics": {}
    }
    
    # Parse "Mécanique et catégories" sheet for definitions
    if "Mécanique et catégories" in wb.sheetnames:
        ws = wb["Mécanique et catégories"]
        print(f"\n=== Parsing 'Mécanique et catégories' sheet ===")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            type_val = row[0]
            name = row[1]
            definition = row[2]
            
            if type_val and name:
                if type_val == "Catégorie":
                    result["categories_definitions"][name] = definition or ""
                    print(f"  Catégorie: {name} -> {definition or 'N/A'}")
                elif type_val == "Mécanique":
                    result["mechanics_definitions"][name] = definition or ""
                    print(f"  Mécanique: {name} -> {definition or 'N/A'}")
    
    # Parse "categories" sheet for game categories
    if "categories" in wb.sheetnames:
        ws = wb["categories"]
        print(f"\n=== Parsing 'categories' sheet ===")
        
        # Get header row (column names)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"  Headers: {headers}")
        
        # Column indices (0-based): Jeux=0, then categories starting from 1
        category_columns = headers[1:]  # Skip "Jeux" column
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            game_name = row[0]
            if not game_name or game_name == "":
                continue
            
            categories = []
            for i, cat_name in enumerate(category_columns):
                col_idx = i + 1  # +1 because row[0] is game name
                if col_idx < len(row) and row[col_idx]:
                    # Check if the cell has a value (1, "1", or "X")
                    cell_value = row[col_idx]
                    if cell_value == 1 or cell_value == "1" or cell_value == "X" or cell_value == "x":
                        categories.append(cat_name)
            
            if categories:
                result["game_categories"][game_name] = categories
                print(f"  {game_name}: {categories}")
    
    # Parse "mecaniques" sheet for game mechanics
    if "mecaniques" in wb.sheetnames:
        ws = wb["mecaniques"]
        print(f"\n=== Parsing 'mecaniques' sheet ===")
        
        # Get header row (column names)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"  Headers: {headers}")
        
        # Column indices (0-based): Jeux=0, then mechanics starting from 1
        mechanic_columns = headers[1:]  # Skip "Jeux" column
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            game_name = row[0]
            if not game_name or game_name == "":
                continue
            
            mechanics = []
            for i, mech_name in enumerate(mechanic_columns):
                col_idx = i + 1  # +1 because row[0] is game name
                if col_idx < len(row) and row[col_idx]:
                    # Check if the cell has a value (X, "X", or 1)
                    cell_value = row[col_idx]
                    if cell_value == "X" or cell_value == "x" or cell_value == 1 or cell_value == "1":
                        mechanics.append(mech_name)
            
            # Store even if empty (to know which games are in the sheet)
            result["game_mechanics"][game_name] = mechanics
            if mechanics:
                print(f"  {game_name}: {mechanics}")
            else:
                print(f"  {game_name}: (aucune mécanique)")
    
    return result


def generate_json_files(data, output_dir):
    """Generate JSON files for each game."""
    
    games_dir = Path(output_dir) / "public" / "data" / "games"
    games_dir.mkdir(parents=True, exist_ok=True)
    
    # Combine all games
    all_games = set(data["game_categories"].keys()) | set(data["game_mechanics"].keys())
    
    print(f"\n=== Generating JSON files for {len(all_games)} games ===")
    
    for game_name in sorted(all_games):
        # Generate file name from game name
        file_name = game_name.lower()
        file_name = file_name.replace("'", "-")
        file_name = file_name.replace(" ", "-")
        file_name = file_name.replace(":", "-")
        file_name = file_name.replace("!", "")
        file_name = file_name.replace("é", "e")
        file_name = file_name.replace("è", "e")
        file_name = file_name.replace("ê", "e")
        file_name = file_name.replace("à", "a")
        file_name = file_name.replace("ù", "u")
        file_name = file_name.replace("ç", "c")
        file_name = file_name.replace("--", "-")
        file_name = file_name.strip("-")
        
        file_path = games_dir / f"{file_name}.json"
        
        # Check if file exists
        if file_path.exists():
            with open(file_path, 'r', encoding='utf-8') as f:
                game_data = json.load(f)
            
            # Update categories and mechanics
            categories = data["game_categories"].get(game_name, [])
            mechanics = data["game_mechanics"].get(game_name, [])
            
            # Apply name transformations
            mechanics = [m.replace("Placement (tuile/dés/ouvries)", "Placements") for m in mechanics]
            mechanics = [m.replace("Développement / engine building", "Développement / Engine building") for m in mechanics]
            mechanics = [m.replace("Alliances (et trahisons)", "Alliances") for m in mechanics]
            
            game_data["categories"] = categories
            game_data["mechanics"] = mechanics
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(game_data, f, ensure_ascii=False, indent=2)
                f.write('\n')
            
            print(f"  Updated: {file_name}.json")
        else:
            print(f"  Not found: {file_name}.json (for game '{game_name}')")


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/parse-excel.py /path/to/Jeux\\ et\\ mécaniques.xlsx")
        print("\nOptions:")
        print("  --json         Output as JSON to stdout")
        print("  --update       Update game JSON files")
        print("  --output DIR   Output directory (default: current directory)")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_json = "--json" in sys.argv
    update_files = "--update" in sys.argv
    
    # Get output directory
    output_dir = "."
    for i, arg in enumerate(sys.argv):
        if arg == "--output" and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    data = parse_excel(excel_path)
    
    if output_json:
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        # Print summary
        print(f"\n=== Summary ===")
        print(f"Categories defined: {len(data['categories_definitions'])}")
        print(f"Mechanics defined: {len(data['mechanics_definitions'])}")
        print(f"Games with categories: {len(data['game_categories'])}")
        print(f"Games with mechanics: {len(data['game_mechanics'])}")
        
        # Print category list
        print(f"\n=== Category list ===")
        for cat in sorted(data['categories_definitions'].keys()):
            print(f"  - {cat}")
        
        # Print mechanic list
        print(f"\n=== Mechanic list ===")
        for mech in sorted(data['mechanics_definitions'].keys()):
            print(f"  - {mech}")
    
    if update_files:
        generate_json_files(data, output_dir)


if __name__ == "__main__":
    main()